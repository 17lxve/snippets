from openpyxl.reader.excel import load_workbook
a = load_workbook("C:/Users/othniel.kouadio/Desktop/smh.xlsx").active
b = load_workbook("C:/Users/othniel.kouadio/Desktop/LDCE.xlsx").active
log = open("log.md",'a')
n=1
s = ""
for row in b.values:
    for value in row:
        s += (value + ",")
    a['A'+str(n)]=s
    n+=1
    log.write(s + "\n")
    s=""